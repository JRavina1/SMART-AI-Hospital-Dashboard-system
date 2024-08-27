import tkinter as tk
from tkinter import messagebox, ttk
from PIL import Image, ImageTk
import os
from openpyxl import load_workbook

# Creating root window
window = tk.Tk()

# Root window title and dimensions
window.title("SMART AI Hospital- Data Files")
window.geometry('700x720')

# Admin credentials
ADMIN_USERNAME = "Admin"
ADMIN_PASSWORD = "Hospital123"

# Path to the Image folder
image_folder = os.path.join(os.path.dirname(__file__), "Image2")

# Data folder path
data_folder = os.path.join(os.path.dirname(__file__), "Data")

# Class for Dashboard
class Dashboard:
    def __init__(self, master):
        self.master = master
        self.current_page = None

        # Create frame for logo
        self.logo_frame = tk.Frame(self.master)
        self.logo_frame.pack(pady=20)

        # Add Hospital logo
        image = Image.open("logo.jpeg")
        image = image.resize((200, 200), Image.LANCZOS)  # Resize the image to 200x200 with LANCZOS resampling
        self.logo_img = ImageTk.PhotoImage(image)  # Creating PhotoImage object from resized image
        self.logo_label = tk.Label(self.logo_frame, image=self.logo_img)
        self.logo_label.pack()

        # Created login frame and adding background color
        self.login_frame = tk.Frame(self.master, background="lightblue")
        self.login_frame.pack(pady=20)

        # Create username label and entry
        self.username_label = tk.Label(self.login_frame, text="Admin Login: ")
        self.username_label.grid(row=0, column=0, padx=15, pady=15)
        self.username_entry = tk.Entry(self.login_frame)
        self.username_entry.grid(row=0, column=1, padx=15, pady=15)

        # Create password label and entry
        self.password_label = tk.Label(self.login_frame, text="Password:")
        self.password_label.grid(row=1, column=0, padx=10, pady=5)
        self.password_entry = tk.Entry(self.login_frame, show="*")
        self.password_entry.grid(row=1, column=1, padx=10, pady=5)

        # Create login button
        self.login_button = tk.Button(self.login_frame, text="Login", command=self.login)
        self.login_button.grid(row=2, columnspan=2, padx=10, pady=5)

    def login(self):
        # Retrieve entered username and password
        username = self.username_entry.get()
        password = self.password_entry.get()

        # Check if the entered credentials match the admin credentials
        if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
            # If correct, show the dashboard
            self.show_dashboard()
        else:
            # If incorrect, show an error message
            messagebox.showerror("Error", "Invalid username or password")

    def show_dashboard(self):
        # Destroy the login frame
        self.login_frame.destroy()

        # Create dashboard frame
        self.dashboard_frame = tk.Frame(self.master)
        self.dashboard_frame.pack(pady=20)
        
        # Create buttons for locking, encrypting, and decrypting files
        self.lock_button = tk.Button(self.dashboard_frame, text="Lock Files", command=self.lock_files)
        self.lock_button.pack()

        self.encrypt_button = tk.Button(self.dashboard_frame, text="Encrypt Files", command=self.encrypt_files)
        self.encrypt_button.pack()

        self.decrypt_button = tk.Button(self.dashboard_frame, text="Decrypt Files", command=self.decrypt_files)
        self.decrypt_button.pack()

        # Create logout button
        self.logout_button = tk.Button(self.dashboard_frame, text="Logout", command=self.logout)
        self.logout_button.pack()

        # Create frame for file data display
        self.file_frame = tk.Frame(self.master)
        self.file_frame.pack(side="right", fill="both", expand=True)

        # Create list of file data with images
        files = [
            ("Staff", "medical-team.png", self.view_staff_file),
            ("Patient", "hospitalization.png", self.view_patient_file),
            ("Department", "Dept.png", self.view_Dept_file),
            ("Hospital", "Hospital.png", self.view_hospital_file),
            ("AI Algorithm", "AIhospital.png", self.view_ai_algorithm_file),
            ("Wards", "wards.png", self.view_wards_file),
            ("Doctor", "doctor.png", self.view_doctor_file),
            ("Pharmacy", "pharmacy.png", self.view_pharmacy_file),
            ("Prescription", "prescription.png", self.view_prescription_file),
            ("Appointment", "schedule.png", self.view_appointment_file),
            ("Invoice", "invoice.png", self.view_invoice_file)
        ]

        for file_name, image_filename, command in files:
            image_path = os.path.join(image_folder, image_filename)
            icon = Image.open(image_path)
            icon = icon.resize((50, 50), Image.LANCZOS)
            icon_img = ImageTk.PhotoImage(icon)
            button = tk.Button(self.file_frame, image=icon_img, text=file_name, compound="top",
                               command=lambda command=command: self.open_page(command))
            button.image = icon_img
            button.pack(side="left", padx=10, pady=10)

    def open_page(self, command):
        # Close the current page if open
        if self.current_page:
            self.current_page.destroy()

        # Create a new page using the provided command
        self.current_page = command()
        self.current_page.pack(fill="both", expand=True)

    def view_staff_file(self):
        # Read the Excel file
        excel_path = os.path.join(data_folder, "Staff_Data.xlsx")
        workbook = load_workbook(excel_path)
        sheet = workbook["Sheet1"]  # Assuming the sheet name is Sheet1

        # Create a new window to display the Excel data
        excel_window = tk.Toplevel(self.master)
        excel_window.title("Staff Data")
        
        # Create a treeview widget to display the data
        tree = ttk.Treeview(excel_window)

        # Define columns and headings
        columns = sheet[1]
        column_names = [column.value for column in columns]
        tree["columns"] = column_names
        for col in column_names:
            tree.column(col, anchor="center", width=100)
            tree.heading(col, text=col, anchor="center")

        # Insert data rows
        for row in sheet.iter_rows(min_row=1, values_only=True):
            tree.insert("", "end", values=row)

        # Pack the treeview widget
        tree.pack(fill="both", expand=True)

    # Define similar functions for other file types
    # (view_patient_file, view_Dept_file, view_hospital_file, view_ai_algorithm_file, view_wards_file, etc.)
    def view_patient_file(self):
        # Read the Excel file
        excel_path = os.path.join(data_folder, "Patient_Data.xlsx")
        workbook = load_workbook(excel_path)
        sheet = workbook["Sheet1"]  # Assuming the sheet name is Sheet1

        # Create a new window to display the Excel data
        excel_window = tk.Toplevel(self.master)
        excel_window.title("Patients Data")

        # Create a treeview widget to display the data
        tree = ttk.Treeview(excel_window)

        # Define columns and headings
        columns = sheet[1]
        column_names = [column.value for column in columns]
        tree["columns"] = column_names
        for col in column_names:
            tree.column(col, anchor="center", width=100)
            tree.heading(col, text=col, anchor="center")

        # Insert data rows
        for row in sheet.iter_rows(min_row=1, values_only=True):
            tree.insert("", "end", values=row)

        # Pack the treeview widget
        tree.pack(fill="both", expand=True)
        
    def view_Dept_file(self):
        # Read the Excel file
        excel_path = os.path.join(data_folder, "Dept_Data.xlsx")
        workbook = load_workbook(excel_path)
        sheet = workbook["Sheet1"]  # Assuming the sheet name is Sheet1

        # Create a new window to display the Excel data
        excel_window = tk.Toplevel(self.master)
        excel_window.title("Department Data")

        # Create a treeview widget to display the data
        tree = ttk.Treeview(excel_window)

        # Define columns and headings
        columns = sheet[1]
        column_names = [column.value for column in columns]
        tree["columns"] = column_names
        for col in column_names:
            tree.column(col, anchor="center", width=100)
            tree.heading(col, text=col, anchor="center")

        # Insert data rows
        for row in sheet.iter_rows(min_row=1, values_only=True):
            tree.insert("", "end", values=row)

        # Pack the treeview widget
        tree.pack(fill="both", expand=True)
        
    def view_doctor_file(self):
        # Read the Excel file
        excel_path = os.path.join(data_folder, "Doctor_Data.xlsx")
        workbook = load_workbook(excel_path)
        sheet = workbook["Sheet1"]  # Assuming the sheet name is Sheet1

        # Create a new window to display the Excel data
        excel_window = tk.Toplevel(self.master)
        excel_window.title("Doctors Data")

        # Create a treeview widget to display the data
        tree = ttk.Treeview(excel_window)

        # Define columns and headings
        columns = sheet[1]
        column_names = [column.value for column in columns]
        tree["columns"] = column_names
        for col in column_names:
            tree.column(col, anchor="center", width=100)
            tree.heading(col, text=col, anchor="center")

        # Insert data rows
        for row in sheet.iter_rows(min_row=1, values_only=True):
            tree.insert("", "end", values=row)

        # Pack the treeview widget
        tree.pack(fill="both", expand=True)    

    def view_hospital_file(self):
        # Read the Excel file
        excel_path = os.path.join(data_folder, "Hospital_Data.xlsx")
        workbook = load_workbook(excel_path)
        sheet = workbook["Sheet1"]  # Assuming the sheet name is Sheet1

        # Create a new window to display the Excel data
        excel_window = tk.Toplevel(self.master)
        excel_window.title("Hospital Data")

        # Create a treeview widget to display the data
        tree = ttk.Treeview(excel_window)

        # Define columns and headings
        columns = sheet[1]
        column_names = [column.value for column in columns]
        tree["columns"] = column_names
        for col in column_names:
            tree.column(col, anchor="center", width=100)
            tree.heading(col, text=col, anchor="center")

        # Insert data rows
        for row in sheet.iter_rows(min_row=1, values_only=True):
            tree.insert("", "end", values=row)

        # Pack the treeview widget
        tree.pack(fill="both", expand=True)
        
    def view_ai_algorithm_file(self):
        # Read the Excel file
        excel_path = os.path.join(data_folder, "AI_Data.xlsx")
        workbook = load_workbook(excel_path)
        sheet = workbook["Sheet1"]  # Assuming the sheet name is Sheet1

        # Create a new window to display the Excel data
        excel_window = tk.Toplevel(self.master)
        excel_window.title("Al Algorithm Data")

        # Create a treeview widget to display the data
        tree = ttk.Treeview(excel_window)

        # Define columns and headings
        columns = sheet[1]
        column_names = [column.value for column in columns]
        tree["columns"] = column_names
        for col in column_names:
            tree.column(col, anchor="center", width=100)
            tree.heading(col, text=col, anchor="center")

        # Insert data rows
        for row in sheet.iter_rows(min_row=1, values_only=True):
            tree.insert("", "end", values=row)

        # Pack the treeview widget
        tree.pack(fill="both", expand=True)
        
    def view_pharmacy_file(self):
        # Read the Excel file
        excel_path = os.path.join(data_folder, "Pharmacy_Data.xlsx")
        workbook = load_workbook(excel_path)
        sheet = workbook["Sheet1"]  # Assuming the sheet name is Sheet1

        # Create a new window to display the Excel data
        excel_window = tk.Toplevel(self.master)
        excel_window.title("Pharmacy Data")

        # Create a treeview widget to display the data
        tree = ttk.Treeview(excel_window)

        # Define columns and headings
        columns = sheet[1]
        column_names = [column.value for column in columns]
        tree["columns"] = column_names
        for col in column_names:
            tree.column(col, anchor="center", width=100)
            tree.heading(col, text=col, anchor="center")

        # Insert data rows
        for row in sheet.iter_rows(min_row=1, values_only=True):
            tree.insert("", "end", values=row)

        # Pack the treeview widget
        tree.pack(fill="both", expand=True)  
        
    def view_prescription_file(self):
        # Read the Excel file
        excel_path = os.path.join(data_folder, "Prescription_Data.xlsx")
        workbook = load_workbook(excel_path)
        sheet = workbook["Sheet1"]  # Assuming the sheet name is Sheet1

        # Create a new window to display the Excel data
        excel_window = tk.Toplevel(self.master)
        excel_window.title("Prescription Data")

        # Create a treeview widget to display the data
        tree = ttk.Treeview(excel_window)

   
        # Pack the treeview widget
        tree.pack(fill="both", expand=True)

    def view_wards_file(self):
        # Read the Excel file
        excel_path = os.path.join(data_folder, "Wards_Data.xlsx")
        workbook = load_workbook(excel_path)
        sheet = workbook["Sheet1"]  # Assuming the sheet name is Sheet1

        # Create a new window to display the Excel data
        excel_window = tk.Toplevel(self.master)
        excel_window.title("Wards Data")

        # Create a treeview widget to display the data
        tree = ttk.Treeview(excel_window)

        # Define columns and headings
        columns = sheet[1]
        column_names = [column.value for column in columns]
        tree["columns"] = column_names
        for col in column_names:
            tree.column(col, anchor="center", width=100)
            tree.heading(col, text=col, anchor="center")

        # Insert data rows
        for row in sheet.iter_rows(min_row=1, values_only=True):
            tree.insert("", "end", values=row)

        # Pack the treeview widget
        tree.pack(fill="both", expand=True)

    def view_invoice_file(self):
        # Read the Excel file
        excel_path = os.path.join(data_folder, "invoice_Data.xlsx")
        workbook = load_workbook(excel_path)
        sheet = workbook["Sheet1"]  # Assuming the sheet name is Sheet1

        # Create a new window to display the Excel data
        excel_window = tk.Toplevel(self.master)
        excel_window.title("invoice Data")

        # Create a treeview widget to display the data
        tree = ttk.Treeview(excel_window)

        # Define columns and headings
        columns = sheet[1]
        column_names = [column.value for column in columns]
        tree["columns"] = column_names
        for col in column_names:
            tree.column(col, anchor="center", width=100)
            tree.heading(col, text=col, anchor="center")

        # Insert data rows
        for row in sheet.iter_rows(min_row=1, values_only=True):
            tree.insert("", "end", values=row)

        # Pack the treeview widget
        tree.pack(fill="both", expand=True)
    
    def view_appointment_file(self):
        # Read the Excel file
        excel_path = os.path.join(data_folder, "Appointment_Data.xlsx")
        workbook = load_workbook(excel_path)
        sheet = workbook["Sheet1"]  # Assuming the sheet name is Sheet1

        # Create a new window to display the Excel data
        excel_window = tk.Toplevel(self.master)
        excel_window.title("Appointment Data")

        # Create a treeview widget to display the data
        tree = ttk.Treeview(excel_window)

        # Define columns and headings
        columns = sheet[1]
        column_names = [column.value for column in columns]
        tree["columns"] = column_names
        for col in column_names:
            tree.column(col, anchor="center", width=100)
            tree.heading(col, text=col, anchor="center")

        # Insert data rows
        for row in sheet.iter_rows(min_row=1, values_only=True):
            tree.insert("", "end", values=row)

        # Pack the treeview widget
        tree.pack(fill="both", expand=True)

    def lock_files(self):
        # Add logic to lock files
        for file_name, _, _ in self.files:
            file_path = os.path.join(data_folder, f"{file_name}_Data.xlsx")
            # Implement file locking mechanism

    def encrypt_files(self):
        # Add logic to encrypt files
        for file_name, _, _ in self.files:
            file_path = os.path.join(data_folder, f"{file_name}_Data.xlsx")
            # Implement file encryption mechanism

    def decrypt_files(self):
        # Add logic to decrypt files
        for file_name, _, _ in self.files:
            file_path = os.path.join(data_folder, f"{file_name}_Data.xlsx")
            # Implement file decryption mechanism
    
    def destroy_dashboard(self):
        # Destroy the dashboard frame and associated widgets
        if hasattr(self, "dashboard_frame"):
            self.dashboard_frame.destroy()
        if hasattr(self, "staff_buttons"):
            for button in self.staff_buttons:
                button.destroy()
                

    def logout(self):
        # Destroy the dashboard frame
       # self.dashboard_frame.destroy()
        # Destroy the dashboard frame
        #self.dashboard_frame.destroy()
        
         # Destroy all frames
        if self.login_frame:
            self.login_frame.destroy()
        if self.dashboard_frame:
            self.dashboard_frame.destroy()
        if self.file_frame:
            self.file_frame.destroy()
        if self.current_page:
            self.current_page.destroy()

        # Recreate the login frame
        self.login_frame = tk.Frame(self.master, background="lightblue")
        self.login_frame.pack(pady=20)

        # Create username label and entry
        self.username_label = tk.Label(self.login_frame, text="Admin Login: ")
        self.username_label.grid(row=0, column=0, padx=15, pady=15)
        self.username_entry = tk.Entry(self.login_frame)
        self.username_entry.grid(row=0, column=1, padx=15, pady=15)

        # Create password label and entry
        self.password_label = tk.Label(self.login_frame, text="Password:")
        self.password_label.grid(row=1, column=0, padx=10, pady=5)
        self.password_entry = tk.Entry(self.login_frame, show="*")
        self.password_entry.grid(row=1, column=1, padx=10, pady=5)

        # Create login button
        self.login_button = tk.Button(self.login_frame, text="Login", command=self.login)
        self.login_button.grid(row=2, columnspan=2, padx=10, pady=5)


# Create an instance of the Dashboard class
app = Dashboard(window)

# Run the main event loop
window.mainloop()
