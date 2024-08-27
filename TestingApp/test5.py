import tkinter as tk
from tkinter import messagebox, ttk
from PIL import Image, ImageTk
import os
from openpyxl import load_workbook
import csv
import hashlib


# Root window title and dimensions
window = tk.Tk()
window.title("SMART AI Hospital- Data Files")
window.geometry('700x720')

# Admin credentials
ADMIN_USERNAME = "Admin"
ADMIN_PASSWORD = "Hospital123"

# Path to the Image folder
image_folder = os.path.join(os.path.dirname(__file__), "Image2")

# Data folder path
data_folder = os.path.join(os.path.dirname(__file__), "Data")
cvs_data_folder = os.path.join(os.path.dirname(__file__), "cvsfiles")
csv_file_path = os.path.join(cvs_data_folder, "Algorithms.csv")



class Dashboard:
    def __init__(self, master):
        self.master = master
        self.current_page = None

        # Create frame for logo
        self.logo_frame = tk.Frame(self.master)
        self.logo_frame.pack(pady=20)

        # Add Hospital logo
        image = Image.open("logo.jpeg")
        image = image.resize((200, 200), Image.LANCZOS)
        self.logo_img = ImageTk.PhotoImage(image)
        self.logo_label = tk.Label(self.logo_frame, image=self.logo_img)
        self.logo_label.pack()

        # Created login frame and adding background color
        self.login_frame = tk.Frame(self.master, background="lightblue")
        self.login_frame.pack(pady=20)

        # Create username label and entry
        self.username_label = tk.Label(self.login_frame, text="Admin Login: ")
        self.username_label.grid(row=0, column=0, padx=15, pady=15)
        self.username_entry = tk.Entry(self.login_frame)
        self.username_entry.grid(row=0, column=1, padx=15, pady=15)

        # Create password label and entry
        self.password_label = tk.Label(self.login_frame, text="Password:")
        self.password_label.grid(row=1, column=0, padx=10, pady=5)
        self.password_entry = tk.Entry(self.login_frame, show="*")
        self.password_entry.grid(row=1, column=1, padx=10, pady=5)

        # Create login button
        self.login_button = tk.Button(self.login_frame, text="Login", command=self.login)
        self.login_button.grid(row=2, columnspan=2, padx=10, pady=5)

    def login(self):
        # Retrieve entered username and password
        username = self.username_entry.get()
        password = self.password_entry.get()

        # Check if the entered credentials match the admin credentials
        if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
            # If correct, show the dashboard
            self.show_dashboard()
        else:
            # If incorrect, show an error message
            messagebox.showerror("Error", "Invalid username or password")

    def show_dashboard(self):
        # Destroy the login frame
        self.login_frame.destroy()

        # Create dashboard frame
        self.dashboard_frame = tk.Frame(self.master)
        self.dashboard_frame.pack(pady=20)

        # Create buttons for locking, unlock, encrypt, and decrypt files
        self.lock_button = tk.Button(self.dashboard_frame, text="Lock Files", command=self.lock_files)
        self.lock_button.pack()

        self.unlock_button = tk.Button(self.dashboard_frame, text="Unlock Files", command=self.unlock_files)
        self.unlock_button.pack()

        self.encrypt_button = tk.Button(self.dashboard_frame, text="Encrypt Files", command=self.encrypt_files)
        self.encrypt_button.pack()

        self.decrypt_button = tk.Button(self.dashboard_frame, text="Decrypt Files", command=self.decrypt_files)
        self.decrypt_button.pack()

        self.directoryCheck_button = tk.Button(self.dashboard_frame, text="Files Directories check", command=self.directoryCheck_files)
        self.directoryCheck_button.pack()
        
        # Create a button to scan directory and create CSV with hashes
        self.scan_button = tk.Button(window, text="Scan Directory and Create CSV", command=lambda: self.create_csv_with_hashes(cvs_data_folder))
        self.scan_button.pack(pady=10)

        # Create a button to compare file hashes
        self.compare_button = tk.Button(window, text="Compare File Hashes", command=self.compare_file_hashes)
        self.compare_button.pack(pady=10)

        # Create logout button
        self.logout_button = tk.Button(self.dashboard_frame, text="Logout", command=self.logout)
        self.logout_button.pack()

        # Create frame for file data display
        self.file_frame = tk.Frame(self.master)
        self.file_frame.pack(side="right", fill="both", expand=True)

        # Create list of file data with images
        files = [
            ("Staff", "medical-team.png", self.view_staff_file),
            ("Patient", "hospitalization.png", self.view_patient_file),
            
        ]

        # Create image buttons
        for file_name, image_filename, command in files:
            image_path = os.path.join(image_folder, image_filename)
            icon = Image.open(image_path)
            icon = icon.resize((50, 50), Image.LANCZOS)
            icon_img = ImageTk.PhotoImage(icon)
            button = tk.Button(self.file_frame, image=icon_img, text=file_name, compound="top",
                               command=lambda command=command: self.open_page(command))
            button.image = icon_img
            button.pack(side="left", padx=10, pady=10)

    def open_page(self, command):
            # Close the current page if open
        if self.current_page:
            self.current_page.destroy()

        # Check if the command is callable
        if callable(command):
            # Call the function and get the returned widget/frame
            page = command()
            if isinstance(page, (tk.Widget, tk.Frame)):
                self.current_page = page
                self.current_page.pack(fill="both", expand=True)
            else:
                messagebox.showerror("Error", "Command did not return a valid page widget.")
        else:
            messagebox.showerror("Error", "Invalid command provided.")

    def view_staff_file(self):
        excel_path = os.path.join(data_folder, "Staff_Data.xlsx")
        lock_file_path = excel_path + ".lock"
        if os.path.exists(lock_file_path):
            messagebox.showinfo("File Locked", "This file is currently locked and cannot be opened.")
            return

        workbook = load_workbook(excel_path)
        sheet = workbook["Sheet1"]

        excel_window = tk.Toplevel(self.master)
        excel_window.title("Staff Data")

        excel_frame = tk.Frame(excel_window)
        excel_frame.pack(fill="both", expand=True)

        tree = ttk.Treeview(excel_frame)

        columns = sheet[1]
        column_names = [column.value for column in columns]
        tree["columns"] = column_names
        for col in column_names:
            tree.column(col, anchor="center", width=100)
            tree.heading(col, text=col, anchor="center")

        # Populate the Treeview with data from the Excel sheet
        for row in sheet.iter_rows(min_row=2, values_only=True):
            tree.insert("", "end", values=row)

        tree.pack(fill="both", expand=True)

        return excel_frame

    def view_patient_file(self):
        excel_path = os.path.join(data_folder, "Patient_Data.xlsx")
        lock_file_path = excel_path + ".lock"
        if os.path.exists(lock_file_path):
            messagebox.showinfo("File Locked", "This file is currently locked and cannot be opened.")
            return

        workbook = load_workbook(excel_path)
        sheet = workbook["Sheet1"]

        excel_window = tk.Toplevel(self.master)
        excel_window.title("Patient Data")

        excel_frame = tk.Frame(excel_window)
        excel_frame.pack(fill="both", expand=True)

        tree = ttk.Treeview(excel_frame)

        columns = sheet[1]
        column_names = [column.value for column in columns]
        tree["columns"] = column_names
        for col in column_names:
            tree.column(col, anchor="center", width=100)
            tree.heading(col, text=col, anchor="center")

        # Populate the Treeview with data from the Excel sheet
        for row in sheet.iter_rows(min_row=2, values_only=True):
            tree.insert("", "end", values=row)

        tree.pack(fill="both", expand=True)

        return excel_frame
    
   

    def lock_files(self):
        image_files = [
            ("Staff", "medical-team.png"),
            ("Patient", "hospitalization.png"),
            
        ]

        for file_name, image_filename in image_files:
            file_path = os.path.join(data_folder, f"{file_name}_Data.xlsx")
            lock_file_path = file_path + ".lock"
            if not os.path.exists(lock_file_path):
                try:
                    with open(lock_file_path, "w") as lock_file:
                        lock_file.write("locked")
                    print(f"File locked successfully: {file_path}")  # Log successful locking
                except Exception as e:
                    print(f"Locking failed for {file_path}: {str(e)}")

    def unlock_files(self):
        image_files = [
            ("Staff", "medical-team.png"),
            ("Patient", "hospitalization.png"),
           
        ]

        for file_name, image_filename in image_files:
            file_path = os.path.join(data_folder, f"{file_name}_Data.xlsx")
            lock_file_path = file_path + ".lock"
            try:
                os.remove(lock_file_path)
                print(f"File unlocked successfully: {file_path}")  # Log successful unlocking
            except FileNotFoundError:
                print(f"File not locked: {file_path}")
            except Exception as e:
                print(f"Unlocking failed for {file_path}: {str(e)}")

    def encrypt_files(self):
        pass

    def decrypt_files(self):
        pass

    def directoryCheck_files(self):
        pass

    # Function to calculate the hash of a file
def calculate_file_hash(file_path):
    with open(file_path, 'rb') as f:
        file_hash = hashlib.sha256()
        while chunk := f.read(4096):
            file_hash.update(chunk)
    return file_hash.hexdigest()

# Function to scan a directory and create a CSV file with hashes for each file
def create_csv_with_hashes(directory):
    with open(csv_file_path, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(['File', 'Hash'])

        for root, _, files in os.walk(directory):
            for file in files:
                file_path = os.path.join(root, file)
                file_hash = calculate_file_hash(file_path)
                writer.writerow([file_path, file_hash])

# Function to compare file hashes with the ones stored in the CSV file
def compare_file_hashes():
    if not os.path.exists(csv_file_path):
        messagebox.showinfo("Error", "CSV file not found. Please create one first.")
        return

    new_file_hashes = {}
    for root, _, files in os.walk(data_folder):
        for file in files:
            file_path = os.path.join(root, file)
            file_hash = calculate_file_hash(file_path)
            new_file_hashes[file_path] = file_hash

    changed_files = []
    with open(csv_file_path, 'r') as csvfile:
        reader = csv.reader(csvfile)
        next(reader)  # Skip header row
        for row in reader:
            file_path, old_hash = row
            if file_path in new_file_hashes:
                new_hash = new_file_hashes[file_path]
                if old_hash != new_hash:
                    changed_files.append(file_path)

    if changed_files:
        messagebox.showinfo("Changed Files", f"The following files have been changed:\n\n{', '.join(changed_files)}")
    else:
        messagebox.showinfo("No Changes", "No files have been changed.")


    def logout(self):
        # Destroy all frames
        if self.login_frame:
            self.login_frame.destroy()
        if self.dashboard_frame:
            self.dashboard_frame.destroy()
        if self.file_frame:
            self.file_frame.destroy()
        if self.current_page:
            self.current_page.destroy()

        # Recreate the login frame
        self.login_frame = tk.Frame(self.master, background="lightblue")
        self.login_frame.pack(pady=20)

        # Create username label and entry
        self.username_label = tk.Label(self.login_frame, text="Admin Login: ")
        self.username_label.grid(row=0, column=0, padx=15, pady=15)
        self.username_entry = tk.Entry(self.login_frame)
        self.username_entry.grid(row=0, column=1, padx=15, pady=15)

        # Create password label and entry
        self.password_label = tk.Label(self.login_frame, text="Password:")
        self.password_label.grid(row=1, column=0, padx=10, pady=5)
        self.password_entry = tk.Entry(self.login_frame, show="*")
        self.password_entry.grid(row=1, column=1, padx=10, pady=5)

        # Create login button
        self.login_button = tk.Button(self.login_frame, text="Login", command=self.login)
        self.login_button.grid(row=2, columnspan=2, padx=10, pady=5)


# Create an instance of the Dashboard class
app = Dashboard(window)

# Run the main event loop
window.mainloop()
